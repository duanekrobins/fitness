import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook

# Define the activity data that an Apple Watch can track
activity_data = {
    'Activity': [
        'Walking', 'Running', 'Cycling', 'Swimming', 'Hiking', 'Strength Training', 'Yoga/Pilates',
        'Rowing', 'Elliptical', 'Skating/Rollerblading', 'Skiing/Snowboarding', 'Dancing', 'Climbing',
        'Martial Arts/Boxing', 'Golf', 'Tennis', 'Horseback Riding', 'Rowing (Water/Erg)'
    ],
    'Metrics Tracked': [
        'Distance, Steps, Active Calories, Heart Rate, Pace',
        'Distance, Steps, Active Calories, Heart Rate, Pace, Cadence',
        'Distance, Active Calories, Heart Rate, Speed, Elevation Gain',
        'Distance, Active Calories, Heart Rate, Strokes, SWOLF',
        'Distance, Steps, Active Calories, Heart Rate, Elevation Gain',
        'Active Calories, Heart Rate, Repetitions, Rest Time',
        'Active Calories, Heart Rate, Mindful Minutes, Flexibility',
        'Distance, Active Calories, Heart Rate, Strokes, Split Time',
        'Distance, Steps, Active Calories, Heart Rate, Strides',
        'Distance, Active Calories, Heart Rate, Speed, Elevation Gain',
        'Distance, Active Calories, Heart Rate, Speed, Elevation Gain',
        'Distance, Steps, Active Calories, Heart Rate, Rhythm',
        'Distance, Active Calories, Heart Rate, Elevation Gain, Speed',
        'Active Calories, Heart Rate, Training Time, Punches, Kicks',
        'Distance, Active Calories, Heart Rate, Strokes, Score',
        'Distance, Active Calories, Heart Rate, Strokes, Score',
        'Distance, Active Calories, Heart Rate, Speed, Elevation Gain',
        'Distance, Active Calories, Heart Rate, Strokes, Split Time'
    ]
}

# Define the health metrics that an Apple Watch can track
health_metrics = {
    'Metric': [
        'Resting Heart Rate', 'Heart Rate Variability (HRV)', 'Blood Pressure', 'Blood Oxygen Level (SpO2)',
        'Body Temperature', 'Respiratory Rate', 'Sweat Rate', 'Glucose Levels', 'Skin Conductance',
        'Body Composition', 'Body Mass Index (BMI)', 'Hydration Level', 'Heart Rate Recovery', 'VO2 Peak', 'Caloric Intake'
    ],
    'Description': [
        'The number of heartbeats per minute when at rest.',
        'The variation in time between consecutive heartbeats.',
        'Systolic and diastolic blood pressure measurements.',
        'The percentage of oxygen in the blood.',
        'Overall body temperature, which can be indicative of various health states.',
        'The number of breaths taken per minute.',
        'Amount of sweat produced, which can indicate hydration levels.',
        'Blood glucose measurements, especially important for diabetics.',
        'Measures electrical conductance of the skin, which can indicate stress levels.',
        'Includes body fat percentage, muscle mass, and bone density.',
        'A measure of body fat based on height and weight.',
        'Indicates how well-hydrated the body is.',
        'The rate at which heart rate returns to baseline after exercise.',
        'The highest value of VO2 achieved during the incremental exercise.',
        'The number of calories consumed.'
    ]
}

# Define the sleep data that an Apple Watch can track
sleep_data = {
    'Metric': [
        'Total Sleep Time', 'Time in Bed', 'Sleep Efficiency', 'Deep Sleep Duration', 'Light Sleep Duration',
        'REM Sleep Duration', 'Wake After Sleep Onset', 'Sleep Latency'
    ],
    'Description': [
        'The total amount of sleep time during the night.',
        'The total amount of time spent in bed.',
        'The percentage of time spent asleep while in bed.',
        'The duration of deep sleep stages.',
        'The duration of light sleep stages.',
        'The duration of REM sleep stages.',
        'The amount of time spent awake after initially falling asleep.',
        'The amount of time it takes to fall asleep.'
    ]
}

# Create DataFrames for the data
activity_df = pd.DataFrame(activity_data)
health_metrics_df = pd.DataFrame(health_metrics)
sleep_data_df = pd.DataFrame(sleep_data)

# Convert the DataFrames to lists of lists for appending to Excel sheets
activity_data_list = activity_df.values.tolist()
health_metrics_data_list = health_metrics_df.values.tolist()
sleep_data_list = sleep_data_df.values.tolist()

# Create a new workbook
workbook = Workbook()

# Add the 'Activity Data' sheet
activity_sheet = workbook.create_sheet('Activity Data')

# Add headers to the sheet
activity_sheet.append(['Activity', 'Metrics Tracked'])

# Append data to the activity sheet
def append_data(sheet, data):
    for row_data in data:
        sheet.append(row_data)

append_data(activity_sheet, activity_data_list)

# Add the 'Health Metrics' sheet
health_metrics_sheet = workbook.create_sheet('Health Metrics')

# Add headers to the sheet
health_metrics_sheet.append(['Metric', 'Description'])

# Append data to the health metrics sheet
append_data(health_metrics_sheet, health_metrics_data_list)

# Add the 'Sleep Data' sheet
sleep_data_sheet = workbook.create_sheet('Sleep Data')

# Add headers to the sheet
sleep_data_sheet.append(['Metric', 'Description'])

# Append data to the sleep data sheet
append_data(sleep_data_sheet, sleep_data_list)

# Remove the default sheet created by openpyxl
if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

# Define the output file name and path
output_file_name = 'Apple_Watch_Activity_Health_Sleep_Data.xlsx'
output_file_path = f'/mnt/data/{output_file_name}'

# Save the workbook with the specified output file name
workbook.save(output_file_path)

print(f"Updated file saved at: {output_file_path}")

# Visualization functions

# Bar plot of number of activities
plt.figure(figsize=(10, 6))
sns.countplot(y='Activity', data=activity_df, order=activity_df['Activity'].value_counts().index)
plt.title('Number of Metrics Tracked per Activity')
plt.xlabel('Number of Metrics')
plt.ylabel('Activity')
plt.tight_layout()
plt.savefig('/mnt/data/Activity_Data_Barplot.png')

# Pie chart of health metrics
plt.figure(figsize=(10, 6))
health_metric_counts = health_metrics_df['Metric'].value_counts()
plt.pie(health_metric_counts, labels=health_metric_counts.index, autopct='%1.1f%%', startangle=140)
plt.title('Distribution of Health Metrics')
plt.tight_layout()
plt.savefig('/mnt/data/Health_Metrics_Piechart.png')

# Bar plot of sleep data
plt.figure(figsize=(10, 6))
sns.countplot(y='Metric', data=sleep_data_df, order=sleep_data_df['Metric'].value_counts().index)
plt.title('Number of Metrics Tracked for Sleep')
plt.xlabel('Number of Metrics')
plt.ylabel('Sleep Metric')
plt.tight_layout()
plt.savefig('/mnt/data/Sleep_Data_Barplot.png')

print("Visualizations saved as 'Activity_Data_Barplot.png', 'Health_Metrics_Piechart.png', and 'Sleep_Data_Barplot.png'")
